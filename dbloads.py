import tkinter
from tkinter import *

from tkinter.ttk import *
from tkinter import messagebox

import os
import json 

from tkinter.filedialog import askdirectory
from tkinter.filedialog import askopenfilename

from openpyxl import load_workbook
from time import time, sleep
from datetime import datetime, timedelta

import subprocess as sp

import threading

import dbaccess as db

class Application(Frame):

    def __init__(self, master):

        self.master = master
        self.main_container = Frame(self.master)

        # define variables
        self.origin = os.getcwd()
        self.source = StringVar()
        self.fileType = IntVar()
        self.tag = StringVar()
        self.annotated = IntVar()
        self.notes = IntVar()
        self.notesSource = StringVar()
        self.ecoSource = StringVar()
        self.gamePlayers = []
        self.matchPlayers = False

        self.games = {}

        # Create main frame
        self.main_container.grid(column=0, row=0, sticky=(N,S,E,W))

        # Set Label styles
        Style().configure("M.TLabel", font="Courier 20 bold", height="20", foreground="blue", anchor="center")
        Style().configure("B.TLabel", font="Verdana 8", background="white", width="18")
        Style().configure("G.TLabel", font="Verdana 8")
        Style().configure("L.TLabel", font="Courier 40 bold", width="4")
        Style().configure("MS.TLabel", font="Verdana 10" )
        Style().configure("S.TLabel", font="Verdana 8" )
        Style().configure("G.TLabel", font="Verdana 8")

        # Set button styles
        Style().configure("B.TButton", font="Verdana 8", relief="ridge")

        # Set check button styles
        Style().configure("B.TCheckbutton", font="Verdana 8")
        Style().configure("B.TRadiobutton", font="Verdana 8")
        Style().configure("O.TLabelframe.Label", font="Verdana 8", foreground="black")

        # Create widgets
        self.sep_a = Separator(self.main_container, orient=HORIZONTAL)
        self.sep_b = Separator(self.main_container, orient=HORIZONTAL)
        self.sep_c = Separator(self.main_container, orient=HORIZONTAL)
        self.sep_d = Separator(self.main_container, orient=HORIZONTAL)
        self.sep_e = Separator(self.main_container, orient=HORIZONTAL)
        self.sep_f = Separator(self.main_container, orient=HORIZONTAL)
        self.sep_g = Separator(self.main_container, orient=HORIZONTAL)
        self.sep_h = Separator(self.main_container, orient=HORIZONTAL)
        self.sep_i = Separator(self.main_container, orient=HORIZONTAL)
        self.mainLabel = Label(self.main_container, text="LOAD GAMES", style="M.TLabel" )

        self.sourceOption = LabelFrame(self.main_container, text=' File ', style="O.TLabelframe")
        self.selectSource = Button(self.sourceOption, text="SOURCE", style="B.TButton", command=self.setSource)
        self.sourceLabel = Label(self.sourceOption, text="None", style="B.TLabel" )

        self.sourceType = LabelFrame(self.main_container, text=' File Type ', style="O.TLabelframe")
        self.srcSheet = Radiobutton(self.sourceType, text="Sheet", style="B.TRadiobutton", variable=self.fileType, value=0)
        self.srcText = Radiobutton(self.sourceType, text=" Text", style="B.TRadiobutton", variable=self.fileType, value=1)
        
        self.load = Button(self.main_container, text="LOAD", style="B.TButton", width=25, command=self.loadCheck)
        self.reset = Button(self.main_container, text="RESET", style="B.TButton", width=25, command=self.resetOptions)
        self.exit = Button(self.main_container, text="EXIT", style="B.TButton", command=root.destroy)

        self.progress_bar = Progressbar(self.main_container, orient="horizontal", mode="indeterminate", maximum=50)

        # Position widgets
        self.mainLabel.grid(row=0, column=0, columnspan=4, padx=5, pady=5, sticky='NSEW')
        
        self.sep_a.grid(row=1, column=0, columnspan=4, padx=5, pady=5, sticky='NSEW')

        self.srcSheet.grid(row=0, column=0, padx=(20,5), pady=(5,10), sticky='NSWE')
        self.srcText.grid(row=0, column=1, padx=(50,40), pady=(5,10), sticky='NSWE')
        self.sourceType.grid(row=2, column=0, columnspan=2, padx=5, pady=5, sticky='NSWE')

        self.selectSource.grid(row=0, column=0, columnspan=1, padx=5, pady=(5,10), sticky='NSWE')
        self.sourceLabel.grid(row=0, column=1, columnspan=1, padx=5, pady=(5,10), sticky='NSWE')
        self.sourceOption.grid(row=2, column=2, columnspan=2, padx=5, pady=5, sticky='NSWE')

        self.sep_e.grid(row=6, column=0, columnspan=4, padx=5, pady=5, sticky='NSEW')

        self.load.grid(row=7, column=0, columnspan=2, padx=5, pady=0, sticky='NSEW')
        self.reset.grid(row=7, column=2, columnspan=2, padx=5, pady=0, sticky='NSEW')
        
        self.sep_f.grid(row=8, column=0, columnspan=4, padx=5, pady=5, sticky='NSEW')
        
        self.exit.grid(row=9, column=0, columnspan=4, padx=5, pady=0, sticky='NSEW')

        self.sep_g.grid(row=10, column=0, columnspan=4, padx=5, pady=5, sticky='NSEW')

        self.progress_bar.grid(row=11, column=0, columnspan=4, padx=5, pady=0, sticky='NSEW')

        self.dataconn = db.databaseConn()

    def displayNotesEntry(self):

        self.popNotes = Toplevel(self.main_container)
        self.popNotes.title("TEXT FILE OPTIONS")

        self.pop_a = Separator(self.popNotes, orient=HORIZONTAL)
        self.pop_b = Separator(self.popNotes, orient=HORIZONTAL)
        self.pop_c = Separator(self.popNotes, orient=HORIZONTAL)

        self.notesOption = LabelFrame(self.popNotes, text=' Notes Source ', style="O.TLabelframe")
        self.selectNotes = Button(self.notesOption, text="SELECT", style="B.TButton", command=self.setNotes)
        self.notesLabel = Label(self.notesOption, text="None", style="B.TLabel" )

        self.ecoOption = LabelFrame(self.popNotes, text=' ECO File ', style="O.TLabelframe")
        self.selectEco = Button(self.ecoOption, text="SELECT", style="B.TButton", command=self.setEco)
        self.ecoLabel = Label(self.ecoOption, text="None", style="B.TLabel" )

        self.tagOption = LabelFrame(self.popNotes, text=' Tag Prefix ', style="O.TLabelframe")
        self.tagPrefix = Entry(self.tagOption, textvariable=self.tag, width="15")
        self.annotation = Checkbutton(self.popNotes, text=" Annotated", style="B.TCheckbutton", variable=self.annotated)
        self.withNotes = Checkbutton(self.popNotes, text=" With notes", style="B.TCheckbutton", variable=self.notes)

        self.proceed = Button(self.popNotes, text="PROCEED", style="B.TButton", command=self.loadTextCheck)
        self.cancel = Button(self.popNotes, text="CANCEL", style="B.TButton", command=self.popNotes.destroy)

        self.selectNotes.grid(row=0, column=0, padx=5, pady=(5,10), sticky='NSW')
        self.notesLabel.grid(row=0, column=0, padx=(100,10), pady=(5,10), sticky='NSW')
        self.notesOption.grid(row=1, column=0, columnspan=2, padx=5, pady=5, sticky='NSW')
        self.withNotes.grid(row=1, column=2, padx=(10,10), pady=(20,10), sticky='NSEW')

        self.pop_a.grid(row=2, column=0, columnspan=4, padx=5, pady=5, sticky="NSEW")  

        self.selectEco.grid(row=0, column=0, padx=5, pady=(5,10), sticky='NSW')
        self.ecoLabel.grid(row=0, column=0, padx=(100,10), pady=(5,10), sticky='NSW')
        self.ecoOption.grid(row=3, column=0, columnspan=2, padx=5, pady=5, sticky='NSW')

        self.tagPrefix.grid(row=0, column=0, columnspan=1, padx=(10,10), pady=(5,10), sticky='NSEW')
        self.tagOption.grid(row=3, column=2, columnspan=1, padx=5, pady=(5,10), sticky='NSEW')
        self.annotation.grid(row=3, column=3, padx=(20,20), pady=(20,10), sticky='NSEW')
        
        self.pop_b.grid(row=4, column=0, columnspan=4, padx=5, pady=5, sticky="NSEW")  

        self.proceed.grid(row=5, column=0, columnspan=2, padx=5, pady=5, sticky="NSEW")  
        self.cancel.grid(row=5, column=2, columnspan=2, padx=5, pady=5, sticky="NSEW")  

        self.pop_c.grid(row=6, column=0, columnspan=4, padx=5, pady=5, sticky="NSEW")  

        ph = 210
        pw = 510

        self.popNotes.maxsize(pw, ph)
        self.popNotes.minsize(pw, ph)

        ws = self.popNotes.winfo_screenwidth()
        hs = self.popNotes.winfo_screenheight()

        x = (ws/2) - (pw/2) - 200
        y = (hs/2) - (ph/2)

        self.popNotes.geometry('%dx%d+%d+%d' % (pw, ph, x, y))

        self.annotated.set(0)
        self.notes.set(0)

    def setSource(self):

        pathname = askopenfilename()
        self.source.set(pathname)

        if pathname:
            try:
                if self.source.get().endswith(".xlsx") or self.source.get().endswith(".xls") or self.source.get().endswith(".txt") or self.source.get().endswith(".pgn"):
                    self.sourceLabel["text"] = self.source.get().split("/")[-1]

                else:
                    messagebox.showerror("Invalid file type", "File type must .xlsx, .xls, .txt or .pgn only.")
                    self.source.set('')
            except:
                pass

    def setNotes(self):

        pathname = askopenfilename()
        self.notesSource.set(pathname)

        if pathname:
            try:
                if self.notesSource.get().endswith(".txt"):
                    self.notesSource.set(pathname)
                    self.notesLabel["text"] = self.notesSource.get().split("/")[-1]

                else:
                    messagebox.showerror(parent=self.popNotes, title="Invalid file type", message="Notes file must be a .txt file type.")

                    self.notesSource.set('')
            except:
                pass

    def setEco(self):

        pathname = askopenfilename()
        self.ecoSource.set(pathname)

        if pathname:
            try:
                if self.ecoSource.get().endswith(".json"):
                    self.ecoLabel["text"] = self.ecoSource.get().split("/")[-1]

                else:
                    messagebox.showerror(parent=self.popNotes, title="Invalid file type", message="ECO file must be a .json file type.")
                    self.ecoSource.set('')
            except:
                pass

    def resetOptions(self):

        self.sourceLabel["text"] = ""
        self.source.set("")

    def loadCheck(self):

        if self.fileType.get() == 0:
            if self.source.get():
                self.loadExcelFile()  
            else:
                messagebox.showerror("Missing source", "Source spreadsheet not selected.")
                return
        else:
            self.displayNotesEntry()

    def loadExcelFile(self):
        
        self.progress_bar.start()

        if self.getDetailsAndInsert():
            messagebox.showinfo("Games loaded", "Games successfully loaded to database")
        else:
            messagebox.showerror("Load error", "Error loading moves from Excel sheet selected")

        self.progress_bar.stop()

    def insertFromExcel(self):

        t = threading.Thread(None, self.getDetailsAndInsert, ())
        t.start()

    def getDetailsAndInsert(self):
        
        wb = load_workbook(self.source.get())

        for ws in wb.worksheets:

            if not ws.title.startswith('Index') and not ws.title.startswith('Template') and not ws.title.startswith('Notes'):

                col_a = ['A', 'D', 'G', 'J', 'M', 'P', 'S', 'V']
                col_b = ['B', 'E', 'H', 'K', 'N', 'Q', 'T', 'W']

                for a, b in zip(col_a, col_b):
                    
                    if ws[a+'1'].value:
                        
                        tag = ws[a+'1'].value
                        
                        opening, variation, white, black, result = self.getIndexDetails(ws[a+'1'].value)

                        if opening:

                            insert_sql = '''
                            insert into game_details 
                            (tag, opening, variation, white, black, result )
                            values (%s, %s, %s, %s, %s, %s)
                            '''
                        
                            game = (tag, opening, variation, white, black, result)

                            try:
                                self.dataconn.execute_insert(insert_sql, game)
                            
                            except Exception as e:
                                print(e)
                                return False
                            
                            white_moves = []
                            black_moves = []

                            for i in range(2,80):
                                idx = i

                                if ws[a+str(idx)].value:
                                    white_moves.append(ws[a+str(idx)].value)
                                else:
                                    break
                                    
                                if ws[b+str(idx)].value:
                                    black_moves.append(ws[b+str(idx)].value)
                                    
                            insert_sql='''
                            insert into game_moves (tag, white_moves, black_moves)
                            values (%s, %s, %s)

                            '''
                            
                            moves = (tag, white_moves, black_moves)
                            
                            try:
                                self.dataconn.execute_insert(insert_sql, moves)

                            except Exception as e:
                                print(e)
                                
                                return False
                        # else:
        return True
    
    def getIndexDetails(self, tag):
        
        wb = load_workbook(self.source.get())
        ws  = wb["Index"]
        
        for row in ws.iter_rows():

            if row[0].value == tag:
                
                op = row[1].value
                
                if row[2].value:
                    vr = row[2].value 
                else:
                    vr = None
                
                res = 0
                if row[5].value == "W":
                    res = 1
                elif row[5].value == "B":
                    res = 2
                    
                return op, '', row[3].value, row[4].value, res
            
        return None, None, None, None, None
    
    def loadTextCheck(self):
        
        if self.source.get() and self.ecoSource.get():
            self.popNotes.destroy()
            self.loadTextFile()
        else:
            messagebox.showerror(parent=self.popNotes, title="Missing files", message="Source sheet and ECO file not selected.")
            return

    def loadTextFile(self):

        self.progress_bar.start()

        self.games = {}

        if self.notes.get():
            if self.notesSource.get():
                self.gamePlayers = self.getGameNames(self.notesSource.get())
                self.matchPlayers = True

        self.processTextFile()

        if self.matchPlayers:
            self.selectGames()

        self.getGameOpening()

        self.loadDatabase()

        messagebox.showinfo(title="Games loaded", message="Games successfully loaded to database")

        self.progress_bar.stop()

    def getGameNames(self, gameNotes):

        name_pairs = []
        ctr = 0
        
        with open(gameNotes, 'r') as infile:

            for line in infile:

                if line.startswith('G:'):
                    ctr += 1
                    names = line.split(':')[1]
                    names = names.split(' - ')

                    white = names[0].strip().lower()
                    black = names[1].strip().lower()

                    pair = [white, black]

                    if pair in name_pairs:
                        pass
                    else:
                        name_pairs.append(pair)

        print(f'Total games found : {ctr}')

        return name_pairs

    def processTextFile(self):

        count = 0
        
        file_path = self.source.get()

        with open(file_path, 'r') as gf:
            
            moves_record = ''
            clean_string = ''

            w_found = False
            b_found = False
            start = False
            winner = 0
                    
            for line in gf:

                if line.startswith('[White '):
                    w_name = line.split('"')[1]
                    w_found = True
                    b_found = False
                    
                if line.startswith('[Black '):
                    b_name = line.split('"')[1]
                    b_found = True

                if line.startswith('[Result'):
                    res = line.split('"')[1]
                    winner = 0
                    if res == '1-0':
                        winner = 1
                    elif res == '0-1':
                        winner = 2

                if line.startswith('[ECO '):
                    eco_code = line.split('"')[1]
                
                if line.startswith('1. '):
                    moves_record = ''
                    clean_string = ''
                    start = True
                
                if start:
                    moves_record += line
                
                    if line.find(' 1-0') > 0 or line.find(' 0-1') > 0 or line.find(' 1/2-1/2') > 0: # 
                        
                        start = False
                        
                        moves_record = moves_record.replace('\n',' ')

                        last_move = self.getLastMove(moves_record)
                        clean_string = self.removeParts(moves_record, '(', ')', last_move)
                        clean_string = self.removeParts(clean_string, '{', '}', last_move)
                        clean_string = clean_string.replace('$2 ', '')
                        clean_string = clean_string.replace('$6 ', '')
                        clean_string = clean_string.replace('$11', '')
                        clean_string = clean_string.replace('$17', '')

                        moves_list, w_moves, b_moves = self.listMoves(clean_string)

                        self.writeToDict(count, eco_code, w_name, b_name, w_moves, b_moves, winner)

                        count += 1

    def getLastMove(self, moves_string):

        ctr = 1

        while True:

            check = str(ctr) + '.'
            try:
                
                if moves_string.index(check) > 0:
                    pass
                    
            except:
                break
                
            ctr += 1
            
        return ctr - 1
        
    def removeParts(self, moves_string, start, end, last):

        clean = ''
        
        inside = 0
        
        for i in moves_string:
            if i == start:

                inside += 1

            if i == end:
                if inside > 0:
                    inside -= 1

            if inside:
                pass
            else:
                if i == end:
                    pass
                else:
                    clean += i
                    
        ctr = 1

        while True:
            chk = ' ' +  str(ctr) + '... '
        
            if clean.find(chk) > 0:
                clean = clean.replace(chk, '') 
        
            ctr += 1
        
            if ctr > last:
                break

        return clean
    
    def listMoves(self, moves_string):

        moves = []
        
        record = moves_string.split('.')

        for rec in record:
            
            if len(rec) > 1:
                
                recs = rec.split()
                
                for r in recs:
                    
                    if r.isnumeric() or rec.isspace():
                        pass
                    elif r.find('-') > 0:
                        if r.find('O-O') == 0 or r.find('O-O-O') == 0:
                            moves.append(r)
                    else:
                        moves.append(r.strip())

            else:
                if rec.isnumeric() or rec.isspace():
                    pass
                else:
                    moves.append(rec.strip())

        # split moves into white or black
        w_moves = []
        b_moves = []
        
        for idx, val in enumerate(moves):
            if idx % 2 == 0:
                w_moves.append(val)
            else:
                b_moves.append(val)

        return moves, w_moves, b_moves
    
    def writeToDict(self, count, eco, white, black, w_moves, b_moves, result):

        self.games[count] = {"eco code": eco, "white": white, "black": black, "white_moves": w_moves, "black_moves": b_moves, "result": result}

    def selectGames(self):

        invalid = []

        for k, v in self.games.items():

            valid = False
            
            for w, b in self.gamePlayers:
                if v["white"].lower().find(w) > -1 and v["black"].lower().find(b) > -1:
                    valid = True

            if valid:
                pass
            else:
                invalid.append(k)

        for idx in invalid:
            del self.games[idx]

    def getGameOpening(self):

        with open(self.ecoSource.get(), 'r') as eco_json:
        
            for data_line in eco_json:
                eco_dict = json.loads(data_line)
                
        eco_json.close()
        
        for k, v in self.games.items():

            self.games[k]['opening'] = eco_dict[v['eco code']][0]['opening']

    def loadDatabase(self):

        count = 1
        
        for k, v in self.games.items():

            tag_count = self.tag.get() + '{:03d}'.format(count)
            opening = v['opening']
            variation = ''
            white = v['white']
            black = v['black']
            result = v['result']
            
            white_moves = v['white_moves']
            black_moves = v['black_moves']

            insert_game = '''
            insert into game_details 
            (tag, opening, variation, white, black, result )
            values (%s, %s, %s, %s, %s, %s)
            '''
        
            game = (tag_count, opening, variation, white, black, result)

            if self.dataconn.execute_insert(insert_game, game):
                pass
            else:
                messagebox.showerror("Insert error", "Error inserting to games details table.")
                return 
            
            insert_moves='''
            insert into game_moves (tag, white_moves, black_moves)
            values (%s, %s, %s)

            '''
            
            moves = (tag_count, white_moves, black_moves)

            if self.dataconn.execute_insert(insert_moves, moves):
                pass
            else:
                messagebox.showerror("Insert error", "Error inserting to games moves table.")
                return 

            count += 1

root = Tk()
root.title("LOAD GAMES")

# Set size

wh = 240
ww = 490

root.resizable(height=False, width=False)

root.minsize(ww, wh)
root.maxsize(ww, wh)

# Position in center screen

ws = root.winfo_screenwidth()
hs = root.winfo_screenheight()

# calculate x and y coordinates for the Tk root window
x = (ws/2) - (ww/2)
y = (hs/2) - (wh/2)

root.geometry('%dx%d+%d+%d' % (ww, wh, x, y))

app = Application(root)

root.mainloop()


